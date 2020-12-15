import openpyxl
excel_path=r"C:\Users\sony\Documents\write_book.xlsx"
new_workbook=openpyxl.Workbook()
#To create new sheet
create_sheet=new_workbook.create_sheet("new1",2)
sheet=new_workbook.active
cell=sheet.cell(2,2)
cell.value="value1"
new_workbook.save(excel_path)
print("Created a xls")