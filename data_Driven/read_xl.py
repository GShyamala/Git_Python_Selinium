import openpyxl

excel_path=r"C:\Users\sony\Documents\new_book.xlsx"
workbook=openpyxl.load_workbook(excel_path)
print((type(workbook)))
sheets=workbook.worksheets
print(type(sheets))
print(sheets)
#Particular sheet
sheet=sheets[0]
# to go active sheet
#sheet=workbook.active
cell=sheet.cell(1,1)
data=cell.value
print(data)
#row size
rows=sheet.max_row
columns=sheet.max_column
#Print all data
for i in range(1,rows+1):
    for j in range(1,columns+1):
        cell=sheet.cell(i,j)
        print(cell.value)